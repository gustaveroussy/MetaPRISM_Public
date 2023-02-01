# -*- coding: utf-8 -*-
"""
@created: Nov 19 2021
@modified: Jan 31 2023
@author: Yoann Pradat

    CentraleSupelec
    MICS laboratory
    9 rue Juliot Curie, Gif-Sur-Yvette, 91190 France

    Institut Gustave Roussy
    Prism Center
    114 rue Edouard Vaillant, Villejuif, 94800 France

Oncoplot-like figure detailing treatment resistances
"""

import argparse
import numpy as np
import openpyxl
import pandas as pd
import re
import sys
import seaborn as sns
from itertools import chain, combinations

# for odds ratio and confidence interval
from scipy.special import expit, logit
from scipy.stats import norm
from statsmodels.tools import add_constant
import statsmodels.api as sm

# pyprism
from pyprism.data import load_table, load_cln, load_bio, split_targeted_therapy_targets, load_colors
from pyprism.util import explode_df

# prettypy
import matplotlib.cm as cm
import matplotlib.lines as mlines
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt

# comut
import palettable
from comut import comut

# DATA PREPRO ==========================================================================================================

def load_drugs_and_alterations(cohort,  alterations, tumor_type, filepath_drug_table, level, col_drug, col_tt, col_id,
                               selection=[("Sample_Type", "DNA_T\\|RNA_T", "regex")]):
    # load data tables
    df_cln = load_cln(cohort)
    df_bio = load_bio(cohort)
    df_drug_table = load_table(filepath_drug_table)
    df_alts = pd.read_table(alterations)

    # rename columns
    df_cln = df_cln.rename(columns={"Project_TCGA_More": col_tt})
    df_cln["Sample_Id_DNA_P"] = df_cln[["Sample_Id_DNA_T", "Sample_Id_DNA_N"]].fillna("NA").apply("_vs_".join, axis=1)

    # choose column of drugs
    if col_drug not in df_cln.columns:
        raise ValueError("-ERROR: the level '%s' is not supported" % level)

    # add columns from biospecimen data
    df_cln["Biopsy_Id"] = df_cln["Biopsy_Selected"].apply(lambda x: x.split("|")[0])
    cols_bio = ["Biopsy_Id", "Biopsy_Site", "Biopsy_Subsite", "Study"]
    df_cln = df_cln.merge(df_bio[cols_bio].drop_duplicates(), how="left", on="Biopsy_Id")

    ##### select patients in clinical table

    ## a. general
    if len(selection)>0:
        for col, val, mod in selection:
            if mod=="regex":
                r = re.compile(r'.*(%s).*' % val)
                mask = df_cln[col].apply(lambda x: bool(r.match(x)))
            elif mode=="exact":
                if type(val)!=list:
                    mask = df_cln[col]==val
                else:
                    mask = df_cln[col].isin(val)
            print("-INFO: selected %d/%d patients from values/regex %s of %s" % (sum(mask), len(mask), val, col))
            df_cln = df_cln[mask].copy()

    ## b. tumor type
    if tumor_type != "All":
        tumor_type_split = tumor_type.split("__")
        mask = df_cln[col_tt].isin(tumor_type_split)
        df_cln = df_cln.loc[mask].copy()
        print("-INFO: selected %d/%d patients from tumor types %s" % (sum(mask), len(mask), tumor_type_split))

    ## c. treatments
    mask = ~df_cln[col_drug].isnull()
    df_cln = df_cln.loc[mask].copy()
    print("-INFO: selected %d/%d patients with non-missing treatments data" % (sum(mask), len(mask)))

    #### select patients in alterations table
    samples_dna = df_cln["Sample_Id_DNA_P"].tolist()
    samples_rna = df_cln["Sample_Id_RNA_T"].tolist()
    mask = df_alts["Sample_Id"].isin(samples_dna+samples_rna)
    df_alts = df_alts.loc[mask].copy()
    print("-INFO: selected %d/%d alterations from patients in clinical table" % (sum(mask), len(mask)))

    # if Classes, split by target. Drugs having multiple targets will be counted multiple times
    if level == "Classes":
        df_cln[col_drug] = df_cln[col_drug].apply(split_targeted_therapy_targets)

    # build dataframe of interest
    cols_keep = [col_id, col_tt, "Primary_Site", "Survival_Status", "Survival_Time", col_drug] + cols_bio
    df_drug = df_cln[cols_keep].copy()

    # split aggregated drugs
    df_drug = explode_df(df=df_drug, cols=[col_drug], sep="|")

    # add classes if applicable
    if level == "Drugs":
        cols_lvl = [x for x in df_drug_table if "_Lvl_" in x]
        df_drug_table_unique = df_drug_table[["DCI"]+cols_lvl].drop_duplicates().rename(columns={"DCI": col_drug})
        df_drug = df_drug.merge(df_drug_table_unique, how="left", on=col_drug)

    return df_cln, df_drug, df_alts


def save_table_drugs_exhaustive(df_drug, col_drug, filepath):
    df_drug_all = df_drug.fillna("N/A").groupby([col_drug, "Class_Lvl_1", "Class_Lvl_2", "Class_Lvl_3"]).size()
    df_drug_all = df_drug_all.to_frame("Count").reset_index()
    df_drug_all = df_drug_all.rename(columns={col_drug: "DCI"})
    df_drug_all = df_drug_all.loc[df_drug_all["DCI"]!="None"].copy()
    df_drug_all = df_drug_all.sort_values(by="Count", ascending=False)
    df_drug_all.to_excel(filepath, index=False)
    print("-INFO: table saved at %s" % filepath)


def get_drugs_groupings(tumor_type, filepath_drug_rules):
    wb = openpyxl.load_workbook(filepath_drug_rules)
    tts = wb.sheetnames
    tumor_types = tumor_type.split("__")
    drugs_groupings = {}

    classes_plot_exception = ["ALK INHIB. 3 GEN.", "AR INHIB. 3 GEN."]

    for tt in tts:
        if tt in tumor_types:
            df_rules = pd.read_excel(filepath_drug_rules, sheet_name=tt)
            if df_rules["Hide"].isnull().mean()<1:
                df_rules = df_rules.loc[df_rules["Hide"].str.lower()!="yes"].copy()
            for class_plot in df_rules["Class_Plot"].unique():
                drugs_class = df_rules.loc[df_rules["Class_Plot"]==class_plot, "DCI"].tolist()
                if len(drugs_class)>1 or class_plot in classes_plot_exception:
                    drugs_groupings[class_plot] = drugs_class

    return drugs_groupings


def _replace_old_new(x, old, new):
    if type(x)==str:
        return x.replace(old,new)
    else:
        return x


def apply_drugs_groupings(df_drug, df_alts, col_drug, drugs_groupings):
    for new, olds in drugs_groupings.items():
        print("-INFO: grouping %s into %s" % (olds, new))
        for old in olds:
            df_drug[col_drug] = df_drug[col_drug].replace(to_replace=old, value=new)
            for db in ["Oncokb", "Civic", "LF"]:
                for direction in ["Res", "Sen"]:
                    col_db_drug = "%s_%s_Drug" % (db, direction)
                    if col_db_drug in df_alts:
                        df_alts[col_db_drug] = df_alts[col_db_drug].apply(_replace_old_new, old=old, new=new)
    df_drug = df_drug.drop_duplicates()
    return df_drug, df_alts


def _get_drugs_orderings(tumor_type):
    if tumor_type == "BRCA":
        drugs_orderings = ["Hormonotherapy"]
    else:
        drugs_orderings = []

    return drugs_orderings


def _get_drugs_ignore(tumor_type, filepath_drug_rules):
    wb = openpyxl.load_workbook(filepath_drug_rules)
    tts = wb.sheetnames
    tumor_types = tumor_type.split("__")
    drugs_ignore = []

    for tt in tts:
        if tt in tumor_types:
            df_rules = pd.read_excel(filepath_drug_rules, sheet_name=tt)
            if df_rules["Hide"].isnull().mean()<1:
                df_hide = df_rules.loc[df_rules["Hide"].str.lower()=="yes"].copy()
                drugs_ignore += df_hide["DCI"].tolist()

    return drugs_ignore


def _get_drugs_enforce(tumor_type, filepath_drug_rules):
    wb = openpyxl.load_workbook(filepath_drug_rules)
    tts = wb.sheetnames
    tumor_types = tumor_type.split("__")
    drugs_enforce = []

    for tt in tts:
        if tt in tumor_types:
            df_rules = pd.read_excel(filepath_drug_rules, sheet_name=tt)
            if df_rules["Show"].isnull().mean()<1:
                df_show = df_rules.loc[df_rules["Show"].str.lower()=="yes"].copy()
                df_show_class = df_show.loc[~df_show["Class_Plot"].isnull()]
                drugs_enforce += df_show["DCI"].tolist() + df_show_class["Class_Plot"].tolist()

    return drugs_enforce



def get_ordered_drugs_and_subjects(df_drug, col_drug, df_alts, threshold, tumor_type, filepath_drug_rules):
    subjects = df_drug["Subject_Id"].unique().tolist()

    # duplicates may exist after some groupings of drugs which do not have identical Class levels 1/2/3.
    df_drug_nodup = df_drug[["Subject_Id", col_drug, "Class_Lvl_1"]].drop_duplicates()

    # drugs order
    drugs_orderings = _get_drugs_orderings(tumor_type)
    df_drug_nonone = df_drug_nodup.loc[df_drug[col_drug]!="None"].copy()
    drugs_ordered = df_drug_nonone[col_drug].value_counts().index.tolist()[::-1]
    for drug_class in drugs_orderings:
        df_drug_nonone_class = df_drug_nonone.loc[df_drug_nonone["Class_Lvl_1"]==drug_class]
        drugs_ordered_class = df_drug_nonone_class[col_drug].value_counts().index.tolist()[::-1]
        if len(drugs_ordered_class)>0:
            index_highest = drugs_ordered.index(drugs_ordered_class[-1]) - len(drugs_ordered)
            drugs_ordered = [x for x in drugs_ordered if x not in drugs_ordered_class]
            for i, drug in enumerate(drugs_ordered_class):
                drugs_ordered.insert(index_highest + i + 1, drug)
                index_highest -= 1

    # subjects order using lexicographical ordering
    subjects_no_drug = sorted(list(set(subjects).difference(set(df_drug_nonone["Subject_Id"].tolist()))))
    df_drug_nonone["Received"] = 1
    df_drug_binary = df_drug_nonone.pivot(index=col_drug, columns="Subject_Id", values="Received")
    df_drug_binary = df_drug_binary.loc[drugs_ordered[::-1],:].fillna("0")
    df_drug_binary[df_drug_binary==1] = "1"
    df_drug_string = df_drug_binary.apply(lambda x: "".join(x.tolist()), axis=0)
    subjects_drugs_ordered = df_drug_string.sort_values(ascending=False).index.tolist()
    subjects_ordered = subjects_drugs_ordered + subjects_no_drug

    # keep only drugs received by at least x patients or with annotations of resistance
    drugs_ordered_ann = []
    for drug in drugs_ordered:
        regex_drug = r"(?:(?<=^)|(?<=\|))%s(?=$|\|)" % drug
        mask_drug = pd.Series(False, index=df_alts.index)
        for col in ["Oncokb_Res_Drug", "Civic_Res_Drug", "LF_Res_Drug"]:
            if col in df_alts:
                mask_drug_col = df_alts[col].fillna("N/A").apply(lambda x: bool(re.search(regex_drug, x)))
                mask_drug = mask_drug | mask_drug_col
        if sum(mask_drug) > 0:
            drugs_ordered_ann.append(drug)

    if threshold > 0 and threshold < 1:
        threshold_cnt = threshold*len(subjects_ordered)
    else:
        threshold_cnt = threshold

    print("-selecting drugs received by at least %d patients" % threshold_cnt)
    s_drug_counts = df_drug_nonone[col_drug].value_counts()
    drugs_ordered_min = s_drug_counts[s_drug_counts >= threshold_cnt].index.tolist()

    # enforce or ignore some drugs
    drugs_ignore = _get_drugs_ignore(tumor_type, filepath_drug_rules)
    drugs_enforce = _get_drugs_enforce(tumor_type, filepath_drug_rules)
    drugs_ordered = [x for x in drugs_ordered if x in drugs_ordered_ann or x in drugs_ordered_min or x in drugs_enforce]
    drugs_ordered = [x for x in drugs_ordered if x not in drugs_ignore]

    print("-INFO: %s drugs will be displayed" % len(drugs_ordered))
    print("-INFO: %s subjects will be displayed" % len(subjects_ordered))

    return drugs_ordered, subjects_ordered


def compute_enrichment_and_interval(df_sym, df_dru, samples_all, drugs_all, subcategories_all,
                                    threshold, conf=0.95, metric="probability"):
    if metric not in ["probability", "odds_ratio"]:
        raise ValueError("-only 'probability' and 'odds_ratio' are allowed for metric")

    df_met = pd.DataFrame(columns=["category", "subcategory", "value", "low", "high"])
    if df_sym.shape[0]==0:
        return df_met

    for cat in df_sym["category"].unique():
        df_sym_cat = df_sym.loc[df_sym["category"]==cat].copy()
        df_dru_cat = df_dru.loc[df_dru["category"]==cat].copy()

        for val in subcategories_all:
            df_sym_cat_val = df_sym_cat.loc[df_sym_cat["value"].apply(lambda x: val in x)]
            if df_sym_cat_val.shape[0] >= threshold and df_dru_cat.shape[0] >= threshold:

                sample_cat = df_dru_cat["sample"].unique().tolist()
                sample_val = df_sym_cat_val["sample"].unique().tolist()
                sample_cat_bar = list(set(samples_all).difference(set(sample_cat)))
                sample_val_bar = list(set(samples_all).difference(set(sample_val)))

                x_cat_val = pd.DataFrame({"sample": list(set(sample_cat).intersection(set(sample_val)))})
                x_cat_val[cat] = 1
                x_cat_val[val] = 1

                x_cat_val_bar = pd.DataFrame({"sample": list(set(sample_cat).intersection(set(sample_val_bar)))})
                x_cat_val_bar[cat] = 1
                x_cat_val_bar[val] = 0

                x_cat_bar_val = pd.DataFrame({"sample": list(set(sample_cat_bar).intersection(set(sample_val)))})
                x_cat_bar_val[cat] = 0
                x_cat_bar_val[val] = 1

                x_cat_bar_val_bar = pd.DataFrame({"sample": list(set(sample_cat_bar).intersection(set(sample_val_bar)))})
                x_cat_bar_val_bar[cat] = 0
                x_cat_bar_val_bar[val] = 0

                df = pd.concat((x_cat_val, x_cat_val_bar, x_cat_bar_val, x_cat_bar_val_bar))
                assert df.shape[0] == len(samples_all)

                crosstab = pd.crosstab(df[cat], df[val])

                # all patients that have resistance alteration have received the drug
                # or
                # all patients that have received the drug have a resistance alteration
                if metric=="odds_ratio" and (crosstab.loc[0,1]==0 or crosstab.loc[1,0]==0):
                    # arbitrary
                    value = -1
                    low = -1
                    high = -1
                else:
                    # fit model with cat and intercept only
                    X = add_constant(df[[cat]])
                    y = df[val]

                    model = sm.Logit(y, X).fit(disp=0)
                    X_unique = X.drop_duplicates().sort_values(by=[cat], ascending=True)
                    X_unique = pd.concat((X_unique, (X_unique.iloc[1]-X_unique.iloc[0]).to_frame().T))
                    se = np.sqrt(np.array([xx@model.cov_params()@xx for xx in X_unique.values]))

                    # NOTE: the predict method of model returns expit(X.T %*% beta) that is the odds of 
                    # being val=1 
                    # pred = X.T %*% beta i.e the log of the odds for being val=1
                    # expit(pred) == model.predict(X_unique)
                    # exp(pred) = odds for val=1
                    # exp(pred_[cat=1-cat=0]) = odds ratio for val=1 btw cat=1 and cat=0
                    z_norm = norm.ppf(max((1-conf/2, conf + (1-conf)/2)))
                    pred = np.array([xx@model.params for xx in X_unique.values])

                    if metric=="odds_ratio":
                        value = np.exp(pred[-1])
                        low = np.exp(pred[-1] - z_norm*se[-1])
                        high = np.exp(pred[-1] + z_norm*se[-1])
                    elif metric=="probability":
                        value = expit(pred[-2])
                        low = expit(pred[-2] - z_norm*se[-2])
                        high = expit(pred[-2] + z_norm*se[-2])

                # append row to dataframe
                row = {"category": cat, "subcategory": val, "value": value, "low": low, "high": high}
                df_met = df_met.append(row, ignore_index=True)

    # replace -1 by max value
    mask_na = df_met["value"]==-1
    if sum(mask_na)>0:
        max_met = df_met["value"].max()
        if np.isnan(max_met) or max_met==-1:
            max_met = 10
        size = int(np.floor(np.log10(max_met)))
        max_value = round(max_met*1.5/10**size)*10**size
        df_met.loc[mask_na, "value"] = max_value
        df_met.loc[mask_na, "low"] = max_value
        df_met.loc[mask_na, "high"] = max_value

    # add missing combinations of drugs/subcategories
    drugs_all_rep = drugs_all*len(subcategories_all)
    subcategories_all_rep = [[x] * len(drugs_all) for x in subcategories_all]
    subcategories_all_rep = [item for sublist in subcategories_all_rep for item in sublist]
    df_met_mis = pd.DataFrame({"category": drugs_all_rep, "subcategory": subcategories_all_rep})
    if metric=="probability":
        df_met_mis["value"] = 0
        df_met_mis["low"] = 0
        df_met_mis["high"] = 0
    elif metric=="odds_ratio":
        df_met_mis["value"] = 1
        df_met_mis["low"] = 1
        df_met_mis["high"] = 1

    df_met = pd.concat((df_met, df_met_mis))
    df_met = df_met.drop_duplicates(subset=["category", "subcategory"], keep="first")

    return df_met


def get_tables(df_drug, df_alts, level, col_drug, col_tt, col_id, drugs_ordered, subjects_ordered, threshold_odds,
               threshold_stks, categorical_pairs, tiers_order, tiers_keep=None):
    dfs_data = {}
    df_drug_nonone = df_drug.loc[df_drug[col_drug]!="None"].copy()
    df_drug_inplot = df_drug.loc[df_drug[col_drug].isin(drugs_ordered)]
    cols_comut = ["sample", "category", "value"]

    # dataframe for center comut plot
    df_dru = df_drug_nonone.rename(columns={"Subject_Id": "sample", col_drug: "category"})
    if level=="Classes":
        df_dru["value"] = "Drug received"
    else:
        df_dru["value"] = df_dru["Class_Lvl_1"]
    df_dru = df_dru[cols_comut].drop_duplicates()
    dfs_data["dru"] = df_dru

    # dataframe for categorical data tumor type
    df_typ = df_drug.rename(columns={"Subject_Id": "sample", col_tt: "value"})
    if "typ" in categorical_pairs:
        df_typ["category"] = categorical_pairs["typ"]
    else:
        df_typ["category"] = "Tumor type"
    df_typ = df_typ[cols_comut].drop_duplicates()
    dfs_data["typ"] = df_typ

    # dataframe for categorical data survival status
    df_srv = df_drug.rename(columns={"Subject_Id": "sample", "Survival_Status": "value"})
    if "srv" in categorical_pairs:
        df_srv["category"] = categorical_pairs["srv"]
    else:
        df_srv["category"] = "Survival status"
    df_srv = df_srv[cols_comut].drop_duplicates()
    dfs_data["srv"] = df_srv

    # dataframe for top barplot
    df_bur = df_drug.groupby("Subject_Id")["Class_Lvl_1"].value_counts().unstack(level=-1)
    df_bur = df_bur.fillna(0).astype(int).reset_index()
    df_bur = df_bur.rename(columns={"Subject_Id": "sample"})
    dfs_data["bur"] = df_bur

    # dataframe for side barplot
    df_drug_inplot_nodup = df_drug_inplot[[col_id, col_drug]].drop_duplicates()
    df_frq = df_drug_inplot_nodup[col_drug].value_counts().to_frame("Treated patients").reset_index()
    df_frq = df_frq.rename(columns={"index": "category"})
    dfs_data["frq"] = df_frq

    # dataframe for scatter plot data of symbols and side stacked bar plot of alterations counts
    dfs_sym = []
    dfs_stk = []
    for drug in drugs_ordered:
        dfs_alts_drug = []
        for db in ["Oncokb", "Civic", "LF"]:
            col_db_drug = "%s_Res_Drug" % db
            col_db_level = "%s_Res_Level_Simple" % db
            regex_drug = r"(?:(?<=^)|(?<=\|))%s(?=$|\|)" % drug
            mask_drug = df_alts[col_db_drug].fillna("N/A").apply(lambda x: bool(re.search(regex_drug, x)))
            df_alts_drug = df_alts.loc[mask_drug].dropna(subset=[col_db_level])
            dfs_alts_drug.append(df_alts_drug.rename(columns={"Subject_Id": "sample", col_db_level: "value"}))
        df_alts_drug = pd.concat((dfs_alts_drug))
        df_alts_drug["category"] = drug
        df_alts_drug["Alteration"] = df_alts_drug[["Hugo_Symbol", "Alteration"]].apply(" ".join, axis=1)

        # NOT: in order not to count twice a pair (drug, alteration), each alteration may only contribute once at the 
        # resistance to the drug. Only the best level of annotation is kept
        cols_keep = ["Row_Identifier", "Alteration", "Hugo_Symbol"] + cols_comut
        df_alts_drug = df_alts_drug[cols_keep].drop_duplicates()
        df_alts_drug["value"] = pd.Categorical(df_alts_drug["value"], categories=tiers_order)
        df_alts_drug = df_alts_drug.sort_values(by="value").drop_duplicates(subset=["Row_Identifier"], keep="first")
        df_alts_drug["Col_Name"] = df_alts_drug[["Alteration", "value"]].apply(" ".join, axis=1)

        df_sym = df_alts_drug[cols_comut].sort_values(by="value").drop_duplicates()
        if df_sym.shape[0]>0:
            df_sym = df_sym.groupby(by=["sample","category"])["value"].apply(" & ".join).reset_index()

        # for stacked barplot, consider only statistics for patients that have received the drug
        df_stk = df_alts_drug[["category", "Hugo_Symbol", "sample", "Alteration", "value", "Col_Name"]]
        df_drug_received = dfs_data["dru"].loc[dfs_data["dru"]["category"]==drug]
        df_stk = df_stk.loc[df_stk["sample"].isin(df_drug_received["sample"].tolist())]

        dfs_sym.append(df_sym)
        dfs_stk.append(df_stk)

    df_sym = pd.concat(dfs_sym)
    dfs_data["sym"] = df_sym

    # in order to have a manageable number of stacks, alterations not Tier 1 and seen in less than threshold times
    # are collapsed to the gene name. Also, all alterations from Literature are collapsed to the gene name unless
    # single. If there is already another gene alteration in Tier 1, label "[Gene] Other".
    df_stk_raw = pd.concat(dfs_stk)

    if tiers_keep is None:
        if df_stk_raw["Alteration"].nunique() >= 20:
            tiers_keep = ["Tier1", "Tier2"]
        else:
            tiers_keep = ["Tier1", "Tier2", "Tier3"]

    # keep alterations detail for Tier1 and Tier2 alterations
    alts_keep = []
    for tier in tiers_keep:
        mask_tier = df_stk_raw["Col_Name"].apply(lambda x: x.endswith(tier))
        alts_tier = df_stk_raw.loc[mask_tier]["Alteration"].unique().tolist()
        alts_keep = list(set(alts_keep).union(set(alts_tier)))

    # keep alterations detail for frequent alterations
    df_stk_raw_cnt = df_stk_raw["Alteration"].value_counts()
    alts_thresh = df_stk_raw_cnt.loc[df_stk_raw_cnt >= threshold_stks].index.tolist()
    alts_keep = list(set(alts_keep).union(set(alts_thresh)))
    genes_keep = df_stk_raw.loc[df_stk_raw["Alteration"].isin(alts_keep)]["Hugo_Symbol"].unique().tolist()

    # other alterations are collapsed to the gene name unless unique
    alts_other = list(set(df_stk_raw["Alteration"]).difference(set(alts_keep)))
    df_other = pd.DataFrame({"Alteration": alts_other, "Hugo_Symbol": [x.split(" ")[0] for x in alts_other]})
    df_other["Hugo_Symbol_Keep"] = df_other["Hugo_Symbol"].apply(lambda x: x in genes_keep)
    df_other_a = df_other[df_other["Hugo_Symbol_Keep"]].copy()
    df_other_b = df_other[~df_other["Hugo_Symbol_Keep"]].copy()

    df_other_a["Alteration_New"] = df_other_a["Hugo_Symbol"] + " Other"
    df_other_b["Alteration_New"] = df_other_b["Hugo_Symbol"]

    df_other = pd.concat((df_other_a, df_other_b))
    old2new = {r["Alteration"]: r["Alteration_New"] for _, r in df_other.iterrows()}
    df_stk_raw["Alteration"] = df_stk_raw["Alteration"].replace(old2new)
    df_stk_raw["Col_Name"] = df_stk_raw[["Alteration", "value"]].apply(" ".join, axis=1)

    # if n alterations contribute to the resistance to the same drug
    #  - if only p < n of these alterations are from the best level, each contributes 1/p and the other n-p contribute 0
    #  - if all n are from the same level each contributes 1/n
    df_stk_raw_uni = df_stk_raw[["category", "sample",  "value"]].drop_duplicates()
    df_stk_raw_uni["value"] = pd.Categorical(df_stk_raw_uni["value"], categories=tiers_order)
    df_stk_raw_uni = df_stk_raw_uni.sort_values(by=["category", "sample", "value"])
    df_stk_raw_uni = df_stk_raw_uni.drop_duplicates(subset=["category", "sample"], keep="first")
    df_stk_raw_uni["Keep"] = "Yes"
    df_stk_raw = df_stk_raw.merge(df_stk_raw_uni, how="left", on=["category", "sample", "value"])
    df_stk_raw = df_stk_raw.loc[df_stk_raw["Keep"]=="Yes"]
    del df_stk_raw["Keep"]

    df_repeat = df_stk_raw.groupby(["category", "sample"]).size().to_frame("Repeat").reset_index()
    df_stk_raw = df_stk_raw.merge(df_repeat, how="left", on=["category", "sample"])
    df_stk_raw["Contrib"] = 1/df_stk_raw["Repeat"]

    df_stk = df_stk_raw.groupby(["category", "Col_Name"])["Contrib"].sum().unstack(level=-1).reset_index()

    # add missing drugs
    drugs_mis = list(set(drugs_ordered).difference(set(df_stk["category"])))
    if len(drugs_mis) > 0:
        df_stk = pd.concat((df_stk, pd.DataFrame({"category": drugs_mis})))
    df_stk = df_stk.fillna(0)

    # order columns
    columns = df_stk.columns[1:].tolist()
    cols_tiers = [(col, col.split(" ")[-1]) for col in columns]
    cols_tiers = sorted(cols_tiers, key=lambda x: x[1])
    columns = [col_tier[0] for col_tier in cols_tiers]
    df_stk = df_stk[["category"]+columns]

    # add missing samples in None colum
    n_subjects = len(subjects_ordered)
    df_stk = df_stk.set_index("category")
    df_rec_tot = dfs_data["dru"]["category"].value_counts()
    df_rec_tot = df_rec_tot.loc[drugs_ordered]
    df_stk_tot = df_stk.sum(axis=1)
    df_stk["None"] = df_rec_tot - df_stk_tot
    df_stk = (df_stk.T/df_rec_tot).T
    df_stk.index.name = "category"
    dfs_data["stk"] = df_stk.reset_index(drop=False)

    # split table per tier level to allow for split bar plots
    for i, tier in enumerate(tiers_order):
        cols_stk_i = [x for x in df_stk if x.endswith(tier)]
        df_stk_i = df_stk[cols_stk_i].copy()
        if i==0:
            df_stk_i["None"] = 1 - df_stk_i[cols_stk_i].sum(axis=1)
        else:
            df_stk_i["None"] = dfs_data["stk_%d" % i]["None"] - df_stk_i[cols_stk_i].sum(axis=1)
            df_stk_i.insert(0, "Previous", dfs_data["stk_%d" % i].sum(axis=1) - dfs_data["stk_%d" % i]["None"])
        dfs_data["stk_%d" % (i+1)] = df_stk_i

    for i in range(1, len(tiers_order)+1):
        dfs_data["stk_%d" % i] = dfs_data["stk_%d" % i].reset_index(drop=False)

    # dataframe for side errorbar plot
    dfs_data["err"] = compute_enrichment_and_interval(df_sym=df_sym, df_dru=df_dru, samples_all=subjects_ordered,
                                                      drugs_all=drugs_ordered,
                                                      subcategories_all=tiers_order,
                                                      threshold=threshold_odds, conf=0.95, metric="odds_ratio")


    return dfs_data


def get_mappings(level, dfs_data, tumor_type, colors_stk="manual", tiers_order=[]):
    cmap_tab20 = cm.get_cmap("tab20")
    cmap_tab20b = cm.get_cmap("tab20b")
    cmap_tab10 = cm.get_cmap("tab10")
    cmap_others = [cm.get_cmap("Oranges"), cm.get_cmap("Purples")]

    mappings = {}
    if level=="Classes":
        mappings["dru"] = {"Drug received": "#F49F0A"}
    else:
        mappings["dru"] = load_colors("Class_Lvl_1")
    mappings["typ"] = load_colors("Project_TCGA_More")
    mappings["srv"] = load_colors("Survival_Status")
    mappings["bur"] = mappings["dru"]
    mappings["frq"] = {"Treated patients": "darkgrey"}

    tiers_colors = {"Tier1": "#DD2D4A", "Tier2": "#FFB700", "Tier3": "#E0AAFF", "Literature": '#3F8EFC'}
    s = list(tiers_order)
    tiers_subsets =  list(chain.from_iterable(combinations(s, r) for r in range(1, len(s)+1)))

    # observe that the tier subsets elements are already ordered
    mappings["sym"] = {}
    for tiers_subset in tiers_subsets:
        tiers_subset = list(tiers_subset)
        color_l = tiers_colors[tiers_subset[0]]
        color_h = tiers_colors[tiers_subset[min(len(tiers_subset)-1, 1)]]
        mappings["sym"][" & ".join(tiers_subset)] = (color_l, color_h)

    mappings["err"] = {"Tier1": [-0.2, '#DD2D4A', '#DD2D4A', '-', 1, 'o', 4],
                       "Tier2": [-0.067, '#FFB700', '#FFB700', '-', 1, 'o', 4],
                       "Tier3": [0.067, '#E0AAFF', '#E0AAFF', '-', 1, 'o', 4],
                       "Literature": [0.2, '#3F8EFC', '#3F8EFC', '-', 1, 'o', 4]}

    # colors for stacked barplot are dictated by the part of the colum names (ignore the "Tier" part at the end)
    if colors_stk=="auto":
        columns_stk = [x for x in dfs_data["stk"].columns if x not in ["category", "None"]]
        columns_sim = [col.split("Tier")[0].strip() for col in columns_stk]
        columns_gen = [col.split(" ")[0] for col in columns_stk]
        df_cols_all = pd.DataFrame({"Stk": columns_stk, "Sim": columns_sim, "Gen": columns_gen})
        df_cols = df_cols_all[["Sim", "Gen"]].drop_duplicates()
        n_alterations = df_cols.shape[0]
        n_genes = df_cols["Gen"].nunique()

        if n_alterations < 10:
            cmap = cmap_tab10
        elif n_alterations < 20 and n_genes <= 5:
            cmap = cmap_tab20b
        else:
            cmap = cmap_tab20

        g_tab = 0
        if cmap==cmap_tab20b:
            g_others = 0
            for gen in df_cols["Gen"].unique():
                n_gen = sum(df_cols["Gen"]==gen)
                if n_gen <= 4:
                    colors_gen = []
                    for i in range(g_tab, g_tab+n_gen):
                        color_rgba = list(cmap(i))
                        color_rgba = [round(x*255) for x in color_rgba]
                        color_hex = '#{:02x}{:02x}{:02x}'.format(*tuple(color_rgba))
                        colors_gen.append(color_hex)
                    g_tab = (g_tab + 4) % 20
                else:
                    colors_gen = []
                    for i in range(n_gen):
                        color_rgba = list(cmap_others[g_others](0.5 + 0.499 * i/(n_gen-1)))
                        color_rgba = [round(x*255) for x in color_rgba]
                        color_hex = '#{:02x}{:02x}{:02x}'.format(*tuple(color_rgba))
                        colors_gen.append(color_hex)
                    g_others = (g_others + 1) % len(cmap_others)
                df_cols.loc[df_cols["Gen"]==gen, "Color"] = colors_gen
        else:
            colors = []
            for i in range(n_alterations):
                color_rgba = list(cmap(i))
                color_rgba = [round(x*255) for x in color_rgba]
                color_hex = '#{:02x}{:02x}{:02x}'.format(*tuple(color_rgba))
                colors.append(color_hex)
            df_cols["Color"] = colors

        df_cols_all = df_cols_all.merge(df_cols, how="left", on=["Sim", "Gen"])

        mappings["stk"] = {stk: col for stk, col in zip(df_cols_all["Stk"], df_cols_all["Color"])}
        mappings["stk"]["None"] = "white"
    elif colors_stk=="manual":
        mappings["stk"] = {}
        if "PRAD" in tumor_type:
            mappings["stk"]['AKT1 Literature'] = '#6F96AC' # new
            mappings["stk"]["AR Amp Literature"] = "#FF7B9C"
            mappings["stk"]["AR Other Literature"] = "#f4acb7"
            mappings["stk"]["AR AR-V7 High Tier2"] = "#F5CCE8"
            mappings["stk"]["AR H875Y Tier2"] = "#C880B7"
            mappings["stk"]["AR H875Y Literature"] = "#C880B7"
            mappings["stk"]["AR T878S Tier2"] = "#9F6BA0"
            mappings["stk"]["AR T878S Literature"] = "#9F6BA0"
            mappings["stk"]['BRAF Literature'] = '#ffcfd2' # new
            mappings["stk"]['HRAS Literature'] = '#add7f6' # new
            mappings["stk"]['NF2 Literature'] = '#80D39B' # new
            mappings["stk"]['PIK3CA Literature'] = '#9d4edd'
            mappings["stk"]['PTEN Literature'] = '#4a4e69'
            mappings["stk"]['RB1 Literature'] = '#FFC759'
        if "LUAD" in tumor_type:
            mappings["stk"]['AKT1 Literature'] = '#6F96AC' # new
            mappings["stk"]['ALK G1202R Tier2'] = '#c9184a'
            mappings["stk"]['ALK L1196M Tier2'] = '#f28482'
            mappings["stk"]['ALK Other Tier3'] = '#ff4d6d'
            mappings["stk"]['BRAF Literature'] = '#ffcfd2' # new
            mappings["stk"]['DNMT3A E505* Tier2'] = '#ffd7ba'
            mappings["stk"]['EGFR T790M Tier1'] = '#8ac926'
            mappings["stk"]['EGFR C797G Tier2'] = '#4f772d'
            mappings["stk"]['EGFR C797S Tier2'] = '#31572c'
            mappings["stk"]['EGFR Amp Tier3'] = '#90a955'
            mappings["stk"]['EGFR Amp Literature'] = '#90a955'
            mappings["stk"]['EGFR Other Literature'] = '#31572c' # new
            mappings["stk"]['ERBB2 Literature'] = '#ffbd00'
            mappings["stk"]['ERBB2 Tier3'] = '#ffbd00'
            mappings["stk"]['KIF5B--RET Literature'] = '#80ffdb' # new
            mappings["stk"]['KRAS G12D Tier3'] = '#3f8efc'
            mappings["stk"]['KRAS G12C Tier2'] = '#87bfff'
            mappings["stk"]['KRAS G12C Literature'] = '#87bfff'
            mappings["stk"]['KRAS G12D Tier2'] = '#3f8efc'
            mappings["stk"]['KRAS G12D Literature'] = '#3f8efc'
            mappings["stk"]['KRAS G13D Tier2'] = '#add7f6'
            mappings["stk"]['KRAS Other Literature'] = '#add7f6' # new
            mappings["stk"]['MAP2K1 Literature'] = '#fbf8cc' # new
            mappings["stk"]['MET Other Literature'] = '#e85d04'
            mappings["stk"]['MET Amp Tier2'] = '#e2711d'
            mappings["stk"]['NF1 Literature'] = '#7209b7' # new
            mappings["stk"]['NF2 Literature'] = '#80D39B' # new
            mappings["stk"]['NRAS Literature'] = '#15616d' # new
            mappings["stk"]['PIK3CA E542Q Literature'] = '#c77dff'
            mappings["stk"]['PIK3CA E542Q Tier2'] = '#c77dff'
            mappings["stk"]['PIK3CA H1047R Tier2'] = '#9b5de5'
            mappings["stk"]['PTEN Literature'] = '#4a4e69'
            mappings["stk"]['STK11 Exon 6 Del Tier2'] = '#b08968'
            mappings["stk"]['STK11 Splice_Site Tier2'] = '#ddb892'
            mappings["stk"]['TP53 DNA Binding Domain Tier3'] = '#be5a38'
            mappings["stk"]['STRN--ALK Literature'] = '#f2e9e4'
            mappings["stk"]['TSC1 Literature'] = '#ccdbfd'
        if "BRCA" in tumor_type:
            mappings["stk"]['ABCC3 Amp Tier3'] = '#73d2de'
            mappings["stk"]['AKT1 Literature'] = '#6f96ac' # new
            mappings["stk"]['AKT2 Literature'] = '#e8fcc2' # new
            mappings["stk"]['RSF1 Amp Tier2'] = '#fec89a'
            mappings["stk"]['TP53 DNA Binding Domain Tier2'] = '#be5a38'
            mappings["stk"]['ESR1 Other Literature'] = '#e05780'
            mappings["stk"]['ESR1 Other Tier3'] = '#e05780'
            mappings["stk"]['ESR1 D538G Tier3'] = '#ffc2d4'
            mappings["stk"]['ESR1 Y537C Tier3'] = '#e05780'
            mappings["stk"]['ESR1 Y537N Tier3'] = '#ff7aa2'
            mappings["stk"]['ESR1 Y537S Tier3'] = '#ff9ebb'
            mappings["stk"]['ESR1--CCDC170 Literature'] = '#ff9ebb'
            mappings["stk"]['HRAS Literature'] = '#add7f6' # new
            mappings["stk"]['KRAS Literature'] = '#3f8efc' # new
            mappings["stk"]['NF1 Literature'] = '#7209b7' # new
            mappings["stk"]['PIK3CA E542K Literature'] = '#c77dff'
            mappings["stk"]['PIK3CA E545K Literature'] = '#ac46a1'
            mappings["stk"]['PIK3CA Other Literature'] = '#47126b'
            mappings["stk"]['PTEN Literature'] = '#4a4e69'
            mappings["stk"]['TP53 DNA Binding Domain Tier3'] = '#be5a38'
            mappings["stk"]['TP53 R273C Tier3'] = '#f3c677'
            mappings["stk"]['TSC1 Literature'] = '#ccdbfd'
        if "BLCA" in tumor_type:
            mappings["stk"]['KRAS Literature'] = '#3f8efc'
            mappings["stk"]['PIK3CA Literature'] = '#9d4edd'
            mappings["stk"]['TSC1 Literature'] = '#ccdbfd'
            mappings["stk"]['FGFR3 Literature'] = '#fe5d9f'
        if "HNSC" in tumor_type:
            mappings["stk"]['EGFR Literature'] = '#9ea93f' # new
            mappings["stk"]['NF1 Literature'] = '#7209b7' # new
            mappings["stk"]['PIK3CA Literature'] = '#9d4edd'
            mappings["stk"]['PTEN Literature'] = '#4a4e69'
        if "CHOL" in tumor_type:
            mappings["stk"]['FGFR2 Literature'] = '#ffb2e6'
            mappings["stk"]['PIK3CA Literature'] = '#9d4edd'
            mappings["stk"]['TSC1 Literature'] = '#ccdbfd'
        if "COAD" in tumor_type:
            mappings["stk"]['AKT1 Literature'] = '#6f96ac' # new
            mappings["stk"]['BRAF V600E Tier2'] = '#7B506F' # new
            mappings["stk"]['KRAS G12A Tier3'] = '#90e0ef'
            mappings["stk"]['KRAS G12V Tier3'] = '#023e8a'
            mappings["stk"]['FBXW7 R505C Tier2'] = '#EB5E55' # new
            mappings["stk"]['NRAS Q61K Tier1'] = '#8C7AA9' # new
            mappings["stk"]['SMAD4 Exon 6 Ins Tier2'] = '#FDF0D5' # new

        mappings["stk"]["Previous"] = "lightgrey"
        mappings["stk"]["None"] = "white"

    mappings["stk_edges"] = {"Previous": None, "Tier1": None, "Tier2": None, "Tier3": None, "Literature": None,
                             "None": "lightgrey"}

    return mappings


def get_widths_heights(n_subjects, n_drugs, n_tiers):
    width_left = 8/8
    width_right_1 = 24*0.5/8
    width_right_2 = 24*0.5/8 * n_tiers
    width_middle = (24*5/8)*(n_subjects/140)
    shadow_width_left = 24*0.85/8
    total_width = width_left + width_middle + width_right_1 + width_right_2 + shadow_width_left
    r_width_left = width_left/total_width
    r_width_right_1 = width_right_1/total_width
    r_width_right_2 = width_right_2/total_width
    r_width_middle = width_middle/total_width
    widths = [r_width_left, r_width_middle, r_width_right_1] + [r_width_right_2/n_tiers] * n_tiers

    height_middle = n_drugs * 6/22
    height_top = 0.5
    total_height = height_top + height_middle
    heights = {'Drug classes': height_top}

    return widths, shadow_width_left, total_width, heights, total_height



def draw_comut_plot(dfs_data, mappings, subjects_ordered, drugs_ordered, tiers_order, widths, shadow_width_left,
                    total_width, heights, total_height, categorical_data=[], categorical_names=[], borders=None,
                    height_middle=None, height_top=None, hspace=0.01, wspace=0.05, label_bar_top="Resistant\ndrugs",
                    label_bar_top_fontsize=12, stacked_bar_top=True, stacked_bar_side=True,
                    label_bar_side="Number of samples", label_bar_side_fontsize=14, mode_bar_side="Number",
                    ncol_legend=None, nrow_legend=None, headers_separate_legend=[],
                    bbox_to_anchor_legend=(1,1), ignored_plots=[], ignored_values=["Previous"], labels_orders={},
                    stk_edges_gap=0.01):
    """
    Instantiate a `CoMut` object and populate with data. Graphical parameters are adjusted based on user parameters.


    Parameters
    ----------
    dfs_data: dict
        Dict of dataframes with data for the figure.
    mappings: dict
        Dict of colors. Keys should match keys of `dfs_data`
    subjects_ordered: list
        Ordered samples.
    drugs_ordered: list
        Ordered drugs.
    categorical_data: list
        List of keys of `dfs_data` holding categorical data to be displayed at the top.
    categorical_names: list
        List of the labels that will be displayed for the categorical data.
    ncol_legend: int
        Number of columns for legend.


    Returns
    ----------
    comut_drug: a `CoMut` object
        Populated with data and parameters set for the rendering the figure.
    """

    comut_drug = comut.CoMut()
    comut_drug.samples = subjects_ordered
    comut_drug.add_categorical_data(data=dfs_data["dru"], name='Drugs received', category_order=drugs_ordered,
                                    mapping=mappings["dru"], xtick_show=False, xtick_fontdict={"fontsize": 8},
                                    ytick_style='italic', ytick_fontdict={"fontsize": 14})


    for cat_data, cat_name in zip(categorical_data, categorical_names):
        comut_drug.add_categorical_data(data=dfs_data[cat_data], name=cat_name,
                                       mapping=mappings[cat_data], xtick_show=False,
                                       ytick_style='normal', ytick_fontdict={"fontsize": 14})

    if dfs_data["sym"].shape[0] > 0:
        comut_drug.add_scatter_data(data=dfs_data["sym"], paired_name='Drugs received', name='Resistances',
                                    mapping=mappings["sym"], marker="*", markersize=8)


    comut_drug.add_side_bar_data(data=dfs_data["frq"], paired_name='Drugs received', name="Treated patients",
                                 position="left", mapping=mappings["frq"], xtick_fontdict={"fontsize": 14},
                                 stacked=stacked_bar_side, xlabel=label_bar_side,
                                 xlabel_fontsize=label_bar_side_fontsize, xlabel_rotation=0,
                                 gap_between_groups=0.1)


    if dfs_data["err"].shape[0] > 0:
        comut_drug.add_side_error_data(data=dfs_data["err"], paired_name='Drugs received', name="Odds ratio",
                                       position="right", mapping=mappings["err"], xtick_fontdict={"fontsize": 12},
                                       xlabel="Odds ratio", xlabel_fontsize=14, xlabel_rotation=0)


    for i, tier in enumerate(tiers_order):
        comut_drug.add_side_bar_data(data=dfs_data["stk_%d" % (i+1)], paired_name="Drugs received",
                                     name="Stacked bars %d" % (i+1), stacked=True, position="right",
                                     mapping=mappings["stk"], xtick_fontdict={"fontsize": 14},
                                     xlabel="Resistances\n%s" % tier, xlabel_fontsize=14, xlabel_rotation=0)

    axis_name_last_stk = "Stacked bars %d" % len(tiers_order)

    # render the plot
    r_shadow_width_left = shadow_width_left/total_width

    comut_drug.plot_comut(x_padding=0.04,
                          y_padding=0.04,
                          tri_padding=0.03,
                          figsize=(total_width,total_height),
                          hspace=hspace,
                          wspace=wspace,
                          heights=heights,
                          widths=widths,
                          shadow_width_left=r_shadow_width_left)


    # configure legends
    handles_more = []
    labels_more = []
    titles_more = []

    # legend for stars
    handles = []
    labels = []
    for label, color in mappings["sym"].items():
        if len(label.split("&"))==1:
            handles.append(mlines.Line2D([], [], color=color[0], marker='*', linestyle='None', markersize=10))
            labels.append(label)

    handles_more.append(handles)
    labels_more.append(labels)
    titles_more.append("Resistances")

    # legends for odds ratios
    handles = []
    labels = []
    for label, (_, color, _, _, _, marker, markersize) in mappings["err"].items():
        handles.append(mlines.Line2D([], [], color=color, marker=marker, linestyle='-', markersize=markersize*1.5))
        labels.append(label)
    handles_more.append(handles)
    labels_more.append(labels)
    titles_more.append("Odds ratios")

    # legend for stacked bars
    for i, tier in enumerate(tiers_order):
        lab2col_i = mappings["stk"]
        lab2col_i = {label: color for label, color in lab2col_i.items() if label in dfs_data["stk_%d" % (i+1)]}
        labs_cols = [(label, color) for label, color in lab2col_i.items() if label not in ["None", "Previous"]]
        labs_cols = sorted(labs_cols, key=lambda x: x[0])
        labs_cols.append(("None", mappings["stk_edges"]["None"]))
        labels = [x[0] for x in labs_cols]
        colors = [x[1] for x in labs_cols]
        df_lc = pd.DataFrame({"label": labels, "color": colors})
        df_lc["label"] = df_lc["label"].apply(lambda x: x.split(tier)[0].strip())
        df_lc = df_lc.drop_duplicates(subset=["label"], keep="first")

        handles = [mpatches.Patch(color=color) if label!="None" else \
                   mpatches.Patch(facecolor="white", edgecolor=color) \
                   for label, color in zip(df_lc["label"], df_lc["color"])]
        labels = [label for label in df_lc["label"]]

        handles_more.append(handles)
        labels_more.append(labels)
        titles_more.append("Alterations %s" % tier)

    comut_drug.add_unified_legend(ncol=ncol_legend, nrow=nrow_legend, axis_name=axis_name_last_stk,
                                  ignored_plots=ignored_plots, ignored_values=ignored_values, handles_more=handles_more,
                                  labels_more=labels_more, titles_more=titles_more,
                                  bbox_to_anchor=bbox_to_anchor_legend, labels_orders=labels_orders,
                                  headers_separate=headers_separate_legend)


    # last adjustments
    if dfs_data["err"].shape[0] > 0:
        axis_name = "Odds ratio"
        min_value = dfs_data["err"]["low"].min()
        min_scale = int(np.floor(np.log10(min_value)))
        min_limit = round(np.floor(min_value/10**min_scale)*10**min_scale, 4)
        max_value = dfs_data["err"]["high"].max()
        max_scale = int(np.floor(np.log10(max_value)))
        max_limit = int(round(np.ceil(max_value/10**max_scale)*10**max_scale, 2))


        if min_limit >= 1:
            xticks = [1/max_limit, max_limit]
            if int(max_limit)==max_limit:
                xticklabels = ["%.1g" % (1/max_limit), "%d" % max_limit]
            else:
                xticklabels = ["%.1g" % (1/max_limit), "%.1f" % max_limit]
        elif max_limit <= 1:
            if min_limit==0:
                xticks = [0.1, 1, 10]
            else:
                xticks = [min_limit, 1]
            xticklabels = xticks
        else:
            if min_limit==0:
                max_limit_plot = max_limit
            else:
                max_limit_plot = max(max_limit, 1/min_limit)
            min_limit_plot = 1/max_limit_plot
            xticks = [min_limit_plot, 1, max_limit_plot]
            if max_limit_plot >= 1e3:
                max_scale = int(np.floor(np.log10(max_limit_plot)))
                xticklabels = ["1e-%d" % max_scale, 1,  "1e%d" % max_scale]
            else:
                if int(max_limit_plot)==max_limit_plot:
                    xticklabels = ["%.1g" % min_limit_plot, 1,  "%d" % max_limit_plot]
                else:
                    xticklabels = ["%.1g" % min_limit_plot, 1,  "%.1f" % max_limit_plot]

        comut_drug.axes[axis_name].axvline(1, color = 'black', linestyle = 'dotted', linewidth = 2)
        comut_drug.axes[axis_name].set_xscale('log')
        comut_drug.axes[axis_name].set_xticks(xticks)
        comut_drug.axes[axis_name].set_xticklabels(xticklabels, fontdict={"fontsize": 14})

    # show spines central comut
    axis_name = "Drugs received"
    for loc in ['bottom']:
        comut_drug.axes[axis_name].spines[loc].set_visible(True)

    # calculate the percentage of samples that received the drug, rounding and adding a percent sign
    axis_name = "Drugs received"
    for loc in ['bottom']:
        comut_drug.axes[axis_name].spines[loc].set_visible(True)

    # calculate the percentage of samples with that gene mutated, rounding and adding a percent sign
    axis_name = "Treated patients"
    df_freq = dfs_data["frq"]
    df_freq = df_freq.set_index("category").loc[drugs_ordered].reset_index()
    percentages = (df_freq[axis_name]/len(comut_drug.samples)*100).round(1).astype(str) + '%'

    # set location of yticks
    comut_drug.axes[axis_name].set_yticks(np.arange(0.5, len(drugs_ordered)+0.5))

    # set labels of yticks
    comut_drug.axes[axis_name].set_yticklabels(list(percentages), ha="right")

    # move the ytick labels inside the bar graph
    comut_drug.axes[axis_name].tick_params(axis='y', pad=0, length=0, labelsize=14)

    # Make y axis visible (by default it is not visible)
    comut_drug.axes[axis_name].get_yaxis().set_visible(True)

    # move y axis ticks to the right
    comut_drug.axes[axis_name].yaxis.tick_right()

    # add edges to stacked bar plot
    for i in range(1, len(tiers_order)+1):
        axis_name = "Stacked bars %d" % i
        comut_drug.axes[axis_name].set_xticks([0,1])
        comut_drug.axes[axis_name].set_xlim([0,1])
        comut_drug.axes[axis_name].set_xticklabels(["0","1"], fontdict={"fontsize": 14})
        df_stk = dfs_data["stk_%d" % i].copy()

        # set correct order for bars
        df_stk_indexed = df_stk.set_index('category')
        df_stk_indexed = df_stk_indexed.reindex(comut_drug._plots["Drugs received"]["data"].index)
        df_stk_edges = pd.DataFrame(index=df_stk_indexed.index)

        patterns = ["Previous"] + tiers_order + ["None"]
        for pattern in patterns:
            cols_pattern = [x for x in df_stk_indexed if x.endswith(pattern)]
            df_stk_edges[pattern] = df_stk_indexed[cols_pattern].sum(axis=1)

        y_range = np.arange(0.5, len(df_stk_edges.index))
        cum_bar_df = np.cumsum(df_stk_edges, axis=1)

        # for each bar, calculate bottom and top of bar and plot
        for j in range(len(cum_bar_df.columns)):
            column = cum_bar_df.columns[j]
            color = mappings["stk_edges"][column]
            if color is not None:
                if j == 0:
                    left = None
                    bar_data = cum_bar_df.loc[:, column]
                else:
                    # calculate distance between previous and current column
                    prev_column = cum_bar_df.columns[j-1]
                    bar_data = cum_bar_df.loc[:, column] - cum_bar_df.loc[:, prev_column] - stk_edges_gap

                    # previous column defines the "bottom" of the bars
                    # 0.01 is to allow some gap
                    left = cum_bar_df.loc[:, prev_column] + stk_edges_gap

                # mask not zero
                mask_nz = bar_data!=0
                if sum(mask_nz) > 0:
                    y_range_nz = y_range[mask_nz]
                    bar_data_nz = bar_data[mask_nz]

                    if left is not None:
                        left_nz = left[mask_nz]
                    else:
                        left_nz = None

                    comut_drug.axes[axis_name].barh(y_range_nz, bar_data_nz, align='center', facecolor='none',
                                                    edgecolor=color, lw=1.5, left=left_nz)

    return comut_drug


def main(args):
    # variable names
    col_drug = "%s_Before_Biopsy" % args.level
    col_tt = "Tumor_Type"
    col_id = "Subject_Id"

    # load data
    df_cln, df_drug, df_alts = load_drugs_and_alterations(cohort=args.cohort, alterations=args.alterations,
                                                          tumor_type=args.tumor_type,
                                                          filepath_drug_table=args.drug_table, level=args.level,
                                                          col_drug=col_drug, col_tt=col_tt, col_id=col_id)

    # add res level for LF annotations
    mask_lf = ~df_alts["LF_Res_Drug"].isnull()
    df_alts.loc[mask_lf, "LF_Res_Level_Simple"] = "Literature"

    # for manual annotations of drug classes
    save_table_drugs_exhaustive(df_drug, col_drug, args.output_table)

    # group some drugs into categories
    drugs_groupings = get_drugs_groupings(args.tumor_type, args.drug_rules)
    df_drug, df_alts = apply_drugs_groupings(df_drug, df_alts, col_drug, drugs_groupings)

    # parameters for the plot 
    cols_comut = ["sample", "category", "value"]
    drugs_ordered, subjects_ordered = get_ordered_drugs_and_subjects(df_drug=df_drug,
                                                                     col_drug=col_drug,
                                                                     df_alts=df_alts,
                                                                     threshold=args.threshold_drug,
                                                                     tumor_type=args.tumor_type,
                                                                     filepath_drug_rules=args.drug_rules)

    # data for the plot
    categorical_data = []
    categorical_names = []
    categorical_pairs = {d: n for d,n in zip(categorical_data, categorical_names)}
    tiers_order = ["Tier1", "Tier2", "Tier3", "Literature"]

    if args.tumor_type=="LUAD":
        tiers_keep = ["Tier1", "Tier2"]
    else:
        tiers_keep = ["Tier1", "Tier2", "Tier3"]

    dfs_data = get_tables(df_drug=df_drug, df_alts=df_alts, level=args.level, col_drug=col_drug, col_tt=col_tt,
                          col_id=col_id, drugs_ordered=drugs_ordered, subjects_ordered=subjects_ordered,
                          threshold_odds=args.threshold_odds, threshold_stks=args.threshold_stks,
                          categorical_pairs=categorical_pairs, tiers_order=tiers_order, tiers_keep=tiers_keep)
    mappings = get_mappings(level=args.level, dfs_data=dfs_data, tumor_type=args.tumor_type, colors_stk="manual",
                            tiers_order=tiers_order)

    # categorical lines
    if args.tumor_type == "All" or len(args.tumor_type.split("__")) > 1:
        categorical_data += ["typ"]
        categorical_names += ["Tumor type"]

    # compute widths and heights - calibrated on LUAD with 140 samples.
    widths, shadow_width_left, total_width, heights, total_height = get_widths_heights(n_subjects=len(subjects_ordered),
                                                                                       n_drugs=len(drugs_ordered),
                                                                                       n_tiers=len(tiers_order))

    bbox_to_anchor_legend = (1.05,1.05)
    ignored_plots = ["Drugs received", "Drug classes", "Resistances"]

    if args.tumor_type == "LUAD":
        headers_separate_legend = ["Odds ratios"] + ["Alterations %s" % tier_order for tier_order in tiers_order]
    else:
        headers_separate_legend = ["Alterations %s" % tier_order for tier_order in tiers_order]

    if args.tumor_type in ["PRAD","BRCA"]:
        nrow_legend = 8
    else:
        nrow_legend = None

    if len(subjects_ordered) < 20:
        wspace = 0.3
    else:
        wspace = 0.2

    comut_drug = draw_comut_plot(dfs_data=dfs_data, mappings=mappings, subjects_ordered=subjects_ordered,
                                 drugs_ordered=drugs_ordered, tiers_order=tiers_order,
                                 categorical_data=categorical_data, categorical_names=categorical_names, widths=widths,
                                 shadow_width_left=shadow_width_left, total_width=total_width,
                                 heights=heights, total_height=total_height, hspace=0.05, wspace=wspace,
                                 label_bar_top="Resistant\ndrugs", label_bar_top_fontsize=14, stacked_bar_top=True,
                                 stacked_bar_side=False, label_bar_side="Treated patients", label_bar_side_fontsize=14,
                                 bbox_to_anchor_legend=bbox_to_anchor_legend, ignored_plots=ignored_plots,
                                 headers_separate_legend=headers_separate_legend, nrow_legend=nrow_legend)

    # save figure
    fig = comut_drug.figure
    fig.savefig(args.output_plot, dpi=300, bbox_inches='tight')
    print("-plot saved at %s" % args.output_plot)

    if args.output_plot_paper is not None:
        fig.savefig(args.output_plot_paper, dpi=300, bbox_inches='tight')
        print("-plot saved at %s" % args.output_plot_paper)


# run ==================================================================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Oncoplot-like figure detailing treatment resistances.")
    parser.add_argument('--cohort', type=str, help='Names of the cohort.', default="prism")
    parser.add_argument('--alterations', type=str, help='Path to table of alterations.',
                default="../../../results/treatment_resistances/annotations/aggregated_alterations_prism_all_more.tsv")
    parser.add_argument('--drug_table', type=str, help='Path to table of drugs.',
                        default="../../../data/resources/drug_tables/Table_Drugs_v7.xlsx")
    parser.add_argument('--drug_rules', type=str, help='Path to rules for grouping of drugs.',
                        default="../../../data/resources/drug_tables/Table_Drugs_Groups_Oncoplot_29072022.xlsx")
    parser.add_argument('--tumor_type', type=str, default="PRAD", help='Selection of a subset of the cohort.')
    parser.add_argument('--level', type=str, default="Drugs",
                        help='Chooose between "Classes" and "Drugs"')
    parser.add_argument('--threshold_drug', type=float, default=5,
                        help='For drugs without any alteration providing resistance in the cohort, only these drugs" + \
                        "received by at least this percentage of patients will be displayed.')
    parser.add_argument('--threshold_odds', type=int, default=4,
        help='Only pairs of (drug, tier) with at least this number of annotations will have an odds ratio computed.')
    parser.add_argument('--threshold_stks', type=int, default=5,
        help="Only alterations seen at least this number of times will have a separate stack, other alterations are" + \
              " collapsed to gene name")
    parser.add_argument('--output_table', type=str,  help='Paths to output oncoplot-like.',
                        default="../../../results/treatment_resistances/plots/comuts/table_PRAD_Drugs.xlsx")
    parser.add_argument('--output_plot', type=str,  help='Paths to output oncoplot-like.',
                        default="../../../results/treatment_resistances/plots/comuts/oncoplot_PRAD_Drugs.pdf")
    parser.add_argument('--output_plot_paper', type=str,  help='Paths to output oncoplot-like.',
                        default=None)
    args = parser.parse_args()

    for arg in vars(args):
        print("%s: %s" % (arg, getattr(args, arg)))
    print("\n", end="")

    main(args)
